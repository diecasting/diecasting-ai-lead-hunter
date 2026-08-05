"""Phase 4 Stage 3.5 — bulk lead import engine (CSV / Excel).

The module is split into three layers so the API router stays thin and the
logic is testable offline:

* ``parse_file``     — pure parsing: bytes -> list of ``{model_field: value}``
                       dicts (one per data row), with column-header aliasing.
* ``preview_rows``   — DB-backed dry run: classify every row as
                       valid / duplicate / failed *without writing anything*.
* ``import_rows``    — DB-backed commit: persist valid rows, skipping
                       duplicate companies / websites, and report per-row
                       outcomes as an ``ImportSummary``.

Accepted CSV columns (case/space/hyphen insensitive, aliases allowed):

    company, website, country, industry, materials, manufacturing_process,
    business_type, buying_signal, contact_role, contact_name, contact_email

``company`` is required and maps to the lead ``name``. Imported leads are
tagged with ``lead_source="import"`` so the dashboard can show provenance.
"""
import csv as _csv
import io
from typing import Dict, List, Optional, Set, Tuple

from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.crud import leads as crud
from app.models.lead import CompanyLead

# Value stored in company_leads.lead_source.
DEFAULT_SOURCE = "import"
MANUAL_SOURCE = "manual"

# Column-header aliases recognised during import. Keys are normalised
# (lower-case, spaces/hyphens -> underscores). Values map to model fields.
COLUMN_ALIASES: Dict[str, str] = {
    "company": "name",
    "company_name": "name",
    "name": "name",
    "country": "country",
    "website": "website",
    "url": "website",
    "industry": "industry",
    "materials": "materials",
    "material": "materials",
    "manufacturing_process": "manufacturing_process",
    "process": "manufacturing_process",
    "business_type": "business_type",
    "buying_signal": "buying_signal",
    "signal": "buying_signal",
    "contact_role": "contact_role",
    "role": "contact_role",
    "contact_name": "contact_name",
    "contact_email": "contact_email",
    "email": "contact_email",
}

# The only model fields the importer may set (extra CSV columns are ignored,
# which keeps a stray "notes" / "phone" column from failing the whole row).
ACCEPTED_FIELDS = (
    "name",
    "website",
    "country",
    "industry",
    "materials",
    "manufacturing_process",
    "business_type",
    "buying_signal",
    "contact_role",
    "contact_name",
    "contact_email",
)


# ---------------------------------------------------------------------------
# Response models
# ---------------------------------------------------------------------------
class ImportRowError(BaseModel):
    """Per-row failure/skip detail returned to the dashboard."""

    row: int
    company: Optional[str] = None
    reason: str


class ImportSummary(BaseModel):
    """Summary of a committed bulk import run."""

    total_rows: int = 0
    imported_count: int = 0
    skipped_count: int = 0
    failed_count: int = 0
    error_details: List[ImportRowError] = []


class ImportPreviewRow(BaseModel):
    """One parsed row with its *would-be* import outcome (dry run)."""

    row: int
    company: Optional[str] = None
    website: Optional[str] = None
    status: str  # valid | duplicate | failed
    reason: Optional[str] = None


class ImportPreview(BaseModel):
    """Dry-run result shown in the modal before the user confirms."""

    total_rows: int = 0
    valid_count: int = 0
    duplicate_count: int = 0
    failed_count: int = 0
    rows: List[ImportPreviewRow] = []


# ---------------------------------------------------------------------------
# Parsing (pure, no DB)
# ---------------------------------------------------------------------------
def norm_header(header: str) -> str:
    """Normalise a header cell: lower-case, spaces/hyphens -> underscores."""
    return (header or "").strip().lower().replace(" ", "_").replace("-", "_")


def _build_norm_map(fieldnames: List[str]) -> Dict[str, str]:
    """Map original header -> normalised model field (aliases resolved)."""
    return {
        orig: COLUMN_ALIASES.get(norm_header(orig), norm_header(orig))
        for orig in fieldnames
    }


def _parse_csv(content: bytes) -> List[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = _csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []
    norm_map = _build_norm_map(fieldnames)
    out: List[dict] = []
    for r in reader:
        row = {norm_map[orig]: (r.get(orig) or "").strip() for orig in fieldnames}
        # Drop fully-empty rows (e.g. trailing blank line in the file).
        if any(row.values()):
            out.append(row)
    return out


def _parse_xlsx(content: bytes) -> List[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        return []
    fieldnames = [str(c) if c is not None else "" for c in header_cells]
    norm_map = _build_norm_map(fieldnames)
    out: List[dict] = []
    for r in rows_iter:
        row = {
            norm_map[orig]: ("" if val is None else str(val)).strip()
            for orig, val in zip(fieldnames, r)
        }
        if any(row.values()):
            out.append(row)
    return out


def parse_file(filename: str, content: bytes) -> List[dict]:
    """Parse an uploaded CSV / Excel file into rows of model fields.

    Raises :class:`ValueError` when the file cannot be parsed (the router
    converts that into an HTTP 400).
    """
    lower = (filename or "").lower()
    try:
        if lower.endswith((".xlsx", ".xlsm")):
            return _parse_xlsx(content)
        return _parse_csv(content)
    except Exception as exc:  # pragma: no cover - parser edge cases
        raise ValueError(f"Failed to parse file: {exc}") from exc


# ---------------------------------------------------------------------------
# Row classification (shared by preview and commit)
# ---------------------------------------------------------------------------
def _classify_row(
    row: dict,
    idx: int,
    existing_names: Set[str],
    existing_websites: Set[str],
    batch_names: Set[str],
) -> Tuple[str, Optional[str], Optional[dict]]:
    """Return ``(status, reason, model_data)`` for one parsed row.

    ``status`` is one of ``valid`` | ``duplicate`` | ``failed``. Model data is
    only produced for ``valid`` rows and always carries ``lead_source``.
    """
    name = (row.get("name") or "").strip()
    if not name:
        return "failed", "missing company name", None
    if name.lower() in existing_names or name.lower() in batch_names:
        return "duplicate", "duplicate company", None

    website = (row.get("website") or "").strip() or None
    if website and website.lower() in existing_websites:
        return "duplicate", "duplicate website", None

    data: dict = {"name": name, "lead_source": DEFAULT_SOURCE}
    for field in ACCEPTED_FIELDS:
        if field == "name":
            continue
        value = (row.get(field) or "").strip()
        if value:
            data[field] = value
    if website:
        data["website"] = website
    return "valid", None, data


def _existing_lookup_sets(db: Session) -> Tuple[Set[str], Set[str]]:
    """Case-insensitive sets of existing company names and websites."""
    names = {
        row[0].lower()
        for row in db.query(CompanyLead.name)
        .filter(CompanyLead.name.isnot(None))
        .all()
    }
    websites = {
        row[0].lower()
        for row in db.query(CompanyLead.website)
        .filter(CompanyLead.website.isnot(None))
        .all()
    }
    return names, websites


# ---------------------------------------------------------------------------
# Dry-run preview + committed import
# ---------------------------------------------------------------------------
def preview_rows(db: Session, rows: List[dict], *, limit: int = 100) -> ImportPreview:
    """Classify every parsed row without persisting anything.

    Used by ``POST /leads/import/preview`` to power the modal's
    preview-before-confirm step. ``rows`` is capped so the response stays
    small even for very large files.
    """
    existing_names, existing_websites = _existing_lookup_sets(db)
    batch_names: Set[str] = set()
    preview = ImportPreview()
    for idx, row in enumerate(rows, start=2):  # row 1 is the header
        preview.total_rows += 1
        status, reason, data = _classify_row(
            row, idx, existing_names, existing_websites, batch_names
        )
        if status == "valid":
            preview.valid_count += 1
            batch_names.add(data["name"].lower())
        elif status == "duplicate":
            preview.duplicate_count += 1
            website = (row.get("website") or "").strip() or None
            preview.rows.append(
                ImportPreviewRow(
                    row=idx,
                    company=(row.get("name") or "").strip() or None,
                    website=website,
                    status=status,
                    reason=reason,
                )
            )
        else:
            preview.failed_count += 1
            preview.rows.append(
                ImportPreviewRow(row=idx, status=status, reason=reason)
            )
        if len(preview.rows) >= limit:
            # Stop appending detail rows, but keep counting.
            preview.rows = preview.rows[:limit]
    return preview


def import_rows(db: Session, rows: List[dict]) -> ImportSummary:
    """Commit valid rows as new leads and report per-row outcomes.

    Deduplication is case-insensitive and considers both the existing
    database and rows earlier in the same file. A row that fails to persist
    (DB error) is counted as failed and rolled back without affecting the
    other rows.
    """
    existing_names, existing_websites = _existing_lookup_sets(db)
    batch_names: Set[str] = set()
    summary = ImportSummary()
    for idx, row in enumerate(rows, start=2):  # row 1 is the header
        summary.total_rows += 1
        status, reason, data = _classify_row(
            row, idx, existing_names, existing_websites, batch_names
        )
        name = (row.get("name") or "").strip()
        if status == "duplicate":
            summary.skipped_count += 1
            summary.error_details.append(
                ImportRowError(row=idx, company=name or None, reason=reason)
            )
            continue
        if status == "failed":
            summary.failed_count += 1
            summary.error_details.append(
                ImportRowError(row=idx, company=None, reason=reason)
            )
            continue
        try:
            crud.create(db, **data)
            summary.imported_count += 1
            batch_names.add(name.lower())
            website = data.get("website")
            if website:
                existing_websites.add(website.lower())
        except Exception as exc:  # pragma: no cover - DB edge cases
            db.rollback()
            summary.failed_count += 1
            summary.error_details.append(
                ImportRowError(row=idx, company=name, reason=f"db error: {exc}")
            )
    return summary
