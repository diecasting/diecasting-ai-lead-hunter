"""Lead API routes: CRUD plus crawl/ingest, AI-analysis, and search endpoints."""
import csv as _csv
import io
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.ai.analyzer import run_analysis
from app.ai.procurement_signals import analyze_procurement_signals
from app.config import settings
from app.crawler.runner import process_pending
from app.crud import leads as crud
from app.database import get_db
from app.models.lead import CompanyLead
from app.outreach.email_generator import generate_email_from_lead
from app.schemas.lead import CompanyLeadCreate, CompanyLeadRead, CompanyLeadUpdate
from app.schemas.outreach import GenerateEmailRequest, OutreachMessageRead
from app.search.service import SearchService

router = APIRouter(prefix="/leads", tags=["leads"])


@router.get("", response_model=List[CompanyLeadRead])
def list_leads(
    db: Session = Depends(get_db),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    relevant_only: bool = False,
    priority: str = Query(None, description="Filter by sales_priority: HIGH/MEDIUM/LOW"),
):
    query = db.query(CompanyLead)
    if relevant_only:
        query = query.filter(CompanyLead.ai_relevant.is_(True))
    if priority:
        query = query.filter(CompanyLead.sales_priority == priority.upper())
    return (
        query.order_by(CompanyLead.id.desc()).offset(skip).limit(limit).all()
    )


@router.post("", response_model=CompanyLeadRead, status_code=201)
def create_lead(payload: CompanyLeadCreate, db: Session = Depends(get_db)):
    if payload.website and crud.get_by_website(db, payload.website):
        raise HTTPException(status_code=409, detail="Lead with this website already exists")
    return crud.create(db, obj_in=payload)


# ---------------------------------------------------------------------------
# Phase 4: Bulk CSV / Excel lead import
# ---------------------------------------------------------------------------
# Column-header aliases recognised during import. Keys are normalised
# (lower-case, spaces/hyphens -> underscores). Values map to model fields.
_COLUMN_ALIASES = {
    "company": "name",
    "company_name": "name",
    "name": "name",
    "country": "country",
    "website": "website",
    "url": "website",
    "industry": "industry",
    "materials": "materials",
    "material": "materials",
    "manufacturing_process": "manufacturing_process",
    "process": "manufacturing_process",
    "buying_signal": "buying_signal",
    "signal": "buying_signal",
    "contact_role": "contact_role",
    "role": "contact_role",
}


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_").replace("-", "_")


class ImportRowError(BaseModel):
    """Per-row failure/skip detail returned to the dashboard."""

    row: int
    company: Optional[str] = None
    reason: str


class LeadImportResult(BaseModel):
    """Summary of a bulk import run."""

    total: int = 0
    imported: int = 0
    skipped: int = 0
    failed: int = 0
    errors: List[ImportRowError] = []


def _parse_csv(content: bytes) -> List[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = _csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []
    norm_map = {
        orig: _COLUMN_ALIASES.get(_norm_header(orig), _norm_header(orig))
        for orig in fieldnames
    }
    return [{norm_map[orig]: (r.get(orig) or "").strip() for orig in fieldnames} for r in reader]


def _parse_xlsx(content: bytes) -> List[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        return []
    fieldnames = [str(c) if c is not None else "" for c in header_cells]
    norm_map = {
        orig: _COLUMN_ALIASES.get(_norm_header(orig), _norm_header(orig))
        for orig in fieldnames
    }
    out: List[dict] = []
    for r in rows_iter:
        out.append(
            {
                norm_map[orig]: ("" if val is None else str(val)).strip()
                for orig, val in zip(fieldnames, r)
            }
        )
    return out


@router.post("/import", response_model=LeadImportResult, status_code=200)
async def import_leads(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Bulk-import leads from a CSV or Excel (.xlsx) file.

    Expected columns (case/space/hyphen insensitive; aliases allowed):
    company, country, website, industry, materials, manufacturing_process,
    buying_signal, contact_role. ``company`` maps to the lead ``name`` and is
    required. Duplicate companies (by name, case-insensitive) and duplicate
    websites are skipped. Rows missing a company name (or that fail to persist)
    are counted as failed with a reason.
    """
    content = await file.read()
    filename = (file.filename or "").lower()
    try:
        if filename.endswith((".xlsx", ".xlsm")):
            rows = _parse_xlsx(content)
        else:
            rows = _parse_csv(content)
    except Exception as exc:  # pragma: no cover - parsing edge cases
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}")

    existing_names = {
        row[0].lower()
        for row in db.query(CompanyLead.name)
        .filter(CompanyLead.name.isnot(None))
        .all()
    }
    existing_websites = {
        row[0].lower()
        for row in db.query(CompanyLead.website)
        .filter(CompanyLead.website.isnot(None))
        .all()
    }
    batch_names: set = set()

    result = LeadImportResult()
    for idx, row in enumerate(rows, start=2):  # row 1 is the header
        result.total += 1
        name = (row.get("name") or "").strip()
        if not name:
            result.failed += 1
            result.errors.append(
                ImportRowError(row=idx, company=None, reason="missing company name")
            )
            continue
        if name.lower() in existing_names or name.lower() in batch_names:
            result.skipped += 1
            result.errors.append(
                ImportRowError(row=idx, company=name, reason="duplicate company")
            )
            continue
        website = (row.get("website") or "").strip() or None
        if website and website.lower() in existing_websites:
            result.skipped += 1
            result.errors.append(
                ImportRowError(row=idx, company=name, reason="duplicate website")
            )
            continue
        data = {
            "name": name,
            "country": row.get("country") or None,
            "website": website,
            "industry": row.get("industry") or None,
            "materials": row.get("materials") or None,
            "manufacturing_process": row.get("manufacturing_process") or None,
            "buying_signal": row.get("buying_signal") or None,
            "contact_role": row.get("contact_role") or None,
            "source": "csv_import",
        }
        try:
            crud.create(db, **data)
            result.imported += 1
            batch_names.add(name.lower())
            if website:
                existing_websites.add(website.lower())
        except Exception as exc:  # pragma: no cover - DB edge cases
            db.rollback()
            result.failed += 1
            result.errors.append(
                ImportRowError(row=idx, company=name, reason=f"db error: {exc}")
            )

    return result


@router.get("/high-priority", response_model=List[CompanyLeadRead])
def get_high_priority_leads(
    db: Session = Depends(get_db),
    limit: int = Query(50, ge=1, le=500),
):
    """Return the most sales-worthy leads, sorted by score descending.

    A lead qualifies as *high-priority* when ``sales_priority`` is ``HIGH``
    (best need-score ≥ 80). If fewer than ``limit`` HIGH leads exist, the
    remaining slots are filled with MEDIUM leads so the sales team always
    has a actionable shortlist.

    NOTE: declared *before* ``/leads/{lead_id}`` so FastAPI matches the static
    path first and ``GET /leads/high-priority`` does not collide with the
    ``{lead_id: int}`` path parameter (which would raise 422 on a non-int value).
    """
    high = (
        db.query(CompanyLead)
        .filter(CompanyLead.sales_priority == "HIGH")
        .order_by(CompanyLead.ai_score.desc().nullslast(), CompanyLead.id.desc())
        .limit(limit)
        .all()
    )
    remaining = limit - len(high)
    if remaining > 0:
        high_ids = {l.id for l in high}
        medium = (
            db.query(CompanyLead)
            .filter(
                CompanyLead.sales_priority == "MEDIUM",
                CompanyLead.id.notin_(high_ids) if high_ids else True,
            )
            .order_by(CompanyLead.ai_score.desc().nullslast(), CompanyLead.id.desc())
            .limit(remaining)
            .all()
        )
        high.extend(medium)
    return high


@router.get("/{lead_id}", response_model=CompanyLeadRead)
def get_lead(lead_id: int, db: Session = Depends(get_db)):
    obj = crud.get(db, lead_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Lead not found")
    return obj


@router.patch("/{lead_id}", response_model=CompanyLeadRead)
def update_lead(lead_id: int, payload: CompanyLeadUpdate, db: Session = Depends(get_db)):
    obj = crud.get(db, lead_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Lead not found")
    return crud.update(db, db_obj=obj, obj_in=payload)


@router.delete("/{lead_id}", status_code=204)
def delete_lead(lead_id: int, db: Session = Depends(get_db)):
    obj = crud.remove(db, lead_id=lead_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Lead not found")


@router.post("/ingest", response_model=List[CompanyLeadRead], status_code=201)
def ingest_from_web(
    db: Session = Depends(get_db),
    query: str = Query("aluminum die casting manufacturer"),
    country: str = Query("us"),
    max_results: int = Query(20, ge=1, le=50),
):
    """Search Google for die-casting companies and persist them as leads.

    Creates ``company_leads`` (dedup by homepage) plus a ``crawl_tasks`` row
    each, ready for the crawler / scheduler. Requires Playwright browsers.
    """
    try:
        report = SearchService().run_search(
            db, query, country=country, max_results=max_results
        )
    except Exception as exc:  # pragma: no cover - depends on network/browser
        raise HTTPException(status_code=502, detail=f"Search failed: {exc}")

    leads = [crud.get(db, lid) for lid in report["created_lead_ids"]]
    leads = [l for l in leads if l is not None]
    return leads


@router.post("/{lead_id}/analyze", response_model=CompanyLeadRead)
def analyze_lead(lead_id: int, db: Session = Depends(get_db)):
    """Run the AI analysis / casting-need scoring on a single lead.

    Works without an OpenAI key (deterministic rule-based scoring). When a key
    is configured, the English summary is enriched by the LLM.
    """
    obj = crud.get(db, lead_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Lead not found")
    try:
        run_analysis(db, obj, crawled_text=obj.description or "")
    except Exception as exc:  # pragma: no cover - depends on OpenAI
        raise HTTPException(status_code=502, detail=f"AI analysis failed: {exc}")
    db.refresh(obj)
    return obj


@router.post("/crawl/pending", response_model=dict)
def crawl_pending(
    db: Session = Depends(get_db),
    limit: int = Query(50, ge=1, le=500),
):
    """Crawl all currently ``pending`` crawl tasks (or use the scheduler)."""
    try:
        from app.crawler.website_crawler import WebsiteCrawler

        report = process_pending(db, limit=limit, crawler=WebsiteCrawler())
    except Exception as exc:  # pragma: no cover - depends on browser/network
        raise HTTPException(status_code=502, detail=f"Crawl failed: {exc}")
    return report


# ---------------------------------------------------------------------------
# Phase 2.3: Industrial AI Lead Intelligence
# ---------------------------------------------------------------------------
@router.post("/{lead_id}/intelligence", response_model=CompanyLeadRead)
def run_intelligence(
    lead_id: int,
    db: Session = Depends(get_db),
    crawl: bool = Query(True, description="Crawl website content before analysis"),
    extract_pdfs: bool = Query(
        True, description="Discover and extract PDF documents (catalogs/brochures)"
    ),
):
    """Full Phase 2.3 intelligence pipeline for a single lead.

    Steps:
    1. (optional) Crawl the company website → structured page text.
    2. (optional) Discover & extract PDF documents → capability signals.
    3. Run the AI analysis (rule-based scoring + optional LLM summary).
    4. Compute the final ``sales_priority`` via the ranking engine.

    The lead's intelligence fields (scores, materials, processes,
    buying_signal, business_type, ai_summary, …) are updated in place.
    """
    obj = crud.get(db, lead_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Lead not found")

    pdf_text = ""

    # Step 1: crawl website content (if enabled and website exists).
    if crawl and obj.website:
        try:
            from app.crawler.website_crawler import WebsiteCrawler

            crawler = WebsiteCrawler()
            result = crawler.crawl(obj.website)
            # result is a CrawlResult dataclass with .text_content and .pages_found
            website_text = getattr(result, "text_content", "") or ""
            if website_text:
                obj.website_content = website_text[:50000]
                obj.pages_crawled = getattr(result, "pages_found", 0) or 0
                from datetime import datetime, timezone
                obj.crawl_status = "completed"
                obj.crawl_time = datetime.now(timezone.utc)
                db.add(obj)
                db.commit()
                # Persist structured human contacts into the contacts table.
                try:
                    from app.crawler.contact_extractor import extract_and_persist

                    extract_and_persist(
                        db, obj.id, website_text, site_domain=obj.domain or ""
                    )
                except Exception:
                    # Contact persistence is best-effort.
                    pass
        except Exception as exc:  # pragma: no cover - depends on browser/network
            # Crawl failure should not block the analysis pipeline.
            obj.crawl_status = "failed"
            db.add(obj)
            db.commit()

    # Step 2: extract PDF documents (if enabled).
    if extract_pdfs and obj.website:
        try:
            from app.crawler.pdf_extractor import PDFExtractor, build_default_fetcher

            # Inject a real HTTP fetcher so PDF discovery can probe the site's
            # document paths. Tests bypass this by constructing PDFExtractor
            # with a mock fetcher directly.
            fetcher = build_default_fetcher()
            extractor = PDFExtractor(fetcher=fetcher)
            home_html = obj.website_content or ""
            pdf_result = extractor.extract_for_lead(db, obj, home_html=home_html)
            # Merge PDF text into the analysis input.
            from app.crud import company_documents as doc_crud
            docs = doc_crud.get_by_lead(db, obj.id)
            pdf_text = " ".join(d.content or "" for d in docs)
        except Exception:  # pragma: no cover - depends on network/PDF libs
            pass

    # Step 3 + 4: AI analysis + ranking (+ procurement signals).
    try:
        crawled_text = " ".join(
            t for t in [obj.website_content or "", pdf_text] if t
        )
        run_analysis(db, obj, crawled_text=crawled_text)

        # Phase 3 Stage 2: industrial procurement-signal analysis.
        procurement = analyze_procurement_signals(crawled_text)
        try:
            import json as _json

            existing = {}
            if obj.ai_signals:
                try:
                    existing = _json.loads(obj.ai_signals)
                except Exception:
                    existing = {}
            existing["procurement_signals"] = {
                "score": procurement["procurement_score"],
                "type": procurement["procurement_type"],
                "components": {
                    k: {"score": v["score"], "matched": v["matched"]}
                    for k, v in procurement["components"].items()
                },
            }
            obj.ai_signals = _json.dumps(existing, ensure_ascii=False)
            db.add(obj)
            db.commit()
        except Exception:
            # Procurement enrichment is best-effort.
            pass
    except Exception as exc:  # pragma: no cover - depends on OpenAI
        raise HTTPException(status_code=502, detail=f"AI analysis failed: {exc}")

    db.refresh(obj)
    return obj


# ---------------------------------------------------------------------------
# Phase 2.4: AI Sales Outreach Engine
# ---------------------------------------------------------------------------
@router.post("/{lead_id}/generate-email", response_model=OutreachMessageRead, status_code=201)
def generate_email(
    lead_id: int,
    db: Session = Depends(get_db),
    overrides: GenerateEmailRequest = None,  # type: ignore[assignment]
):
    """Generate a personalised B2B sales outreach email for this lead.

    Uses the lead's AI intelligence fields (industry, materials, processes,
    buying signal) plus an industry-specific template to produce a technically
    grounded email. When OpenAI is configured the text is enriched by the LLM;
    otherwise a deterministic template render is used.

    The generated email is saved as an ``outreach_messages`` row with status
    ``draft`` so the sales team can review before sending.
    """
    from app.crud import outreach as outreach_crud

    lead = crud.get(db, lead_id)
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")

    try:
        result = generate_email_from_lead(db, lead, use_llm=True)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Email generation failed: {exc}")

    subject = result.get("subject", f"Partnership opportunity with {lead.name or 'your company'}")
    body_parts = []
    opening = result.get("opening", "")
    main_body = result.get("body", "")
    cta = result.get("call_to_action", "")
    if opening:
        body_parts.append(opening)
    if main_body:
        body_parts.append(main_body)
    if cta:
        body_parts.append("\n" + cta)
    full_body = "\n\n".join(body_parts)

    # Phase 4 Stage 2: compute the email quality score (0-100) and persist it so
    # the dashboard can display a quality gate per draft.
    quality_score = None
    try:
        from app.outreach.context import build_context_from_lead
        from app.outreach.email_quality import score_email_quality

        ctx = build_context_from_lead(lead, db=db)
        quality = score_email_quality(full_body, ctx)
        quality_score = quality.get("quality")
    except Exception:
        # Quality scoring is best-effort; never block draft creation.
        quality_score = None

    # Phase 4 Stage 3: classify the quality score into a gate decision
    # (ready | review | blocked) stored on the draft so the CRM can filter /
    # block low-quality drafts before a human reviews them.
    gate_status = None
    try:
        from app.outreach.draft_quality_gate import classify_quality_gate

        gate_status = classify_quality_gate(quality_score)
    except Exception:
        gate_status = None

    msg = outreach_crud.create(
        db,
        lead_id=lead.id,
        subject=subject[:500],
        body=full_body,
        contact_role=result.get("contact_role"),
        status="draft",
        quality_score=quality_score,
        quality_gate_status=gate_status,
    )
    return msg


# ---------------------------------------------------------------------------
# Phase 2.5: CRM pipeline
# ---------------------------------------------------------------------------
class LeadStatusUpdate(BaseModel):
    """Payload for updating a lead's sales pipeline status."""

    lead_status: str
    next_followup_date: Optional[datetime] = None


@router.patch("/{lead_id}/status", response_model=CompanyLeadRead)
def update_lead_status(
    lead_id: int,
    payload: LeadStatusUpdate,
    db: Session = Depends(get_db),
):
    """Update a lead's sales pipeline status (e.g. new → qualified → contacted).

    Validates the transition against the CRM state machine
    (``app.outreach.workflow.VALID_TRANSITIONS``).
    """
    from app.outreach.workflow import transition

    lead = crud.get(db, lead_id)
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")

    try:
        transition(lead, payload.lead_status, db=db)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if payload.next_followup_date is not None:
        lead.next_followup_date = payload.next_followup_date
        db.add(lead)
        db.commit()
        db.refresh(lead)

    db.refresh(lead)
    return lead
